import json
import threading
from datetime import datetime
from openpyxl import Workbook, load_workbook
from modules.rs485 import RS485


class DataProcessor:
    def __init__(self, config_path, output_path):
        self.rs485 = None
        self.output_path = output_path
        self.load_config(config_path)
        self.is_recording = False

        # 初始化 Excel 文件
        self.init_excel()

    def load_config(self, config_path):
        try:
            with open(config_path, 'r') as f:
                config = json.load(f)
            self.rs485 = RS485(
                port=config['serial_port'],
                baudrate=config['baudrate'],
                databits=config['databits'],
                parity=config['parity'],
                stopbits=config['stopbits']
            )
        except Exception as e:
            raise RuntimeError(f"加载配置文件失败: {e}")

    def init_excel(self):
        try:
            # 如果文件已存在，加载现有文件；否则创建新文件
            try:
                self.workbook = load_workbook(self.output_path)
                self.sheet = self.workbook.active
            except FileNotFoundError:
                self.workbook = Workbook()
                self.sheet = self.workbook.active
                self.sheet.title = "Data"
                self.sheet.append(["Timestamp", "FlowRate"])  # 添加表头
                self.workbook.save(self.output_path)
        except Exception as e:
            raise RuntimeError(f"初始化 Excel 文件失败: {e}")

    def start_recording(self, callback):
        self.is_recording = True
        threading.Thread(target=self.record_data, args=(callback,), daemon=True).start()

    def stop_recording(self):
        self.is_recording = False

    def record_data(self, callback):
        try:
            while self.is_recording:
                data = self.rs485.read_data()
                if data:
                    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    flow_rate = data['flow_rate']
                    # 保存到 Excel
                    self.sheet.append([timestamp, flow_rate])
                    self.workbook.save(self.output_path)
                    # 回调更新界面
                    callback(timestamp, flow_rate)
        except Exception as e:
            print(f"数据采集过程中出错: {e}")
